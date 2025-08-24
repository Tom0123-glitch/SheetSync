import os
from typing import Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

APP_TITLE = "SheetSync — Excel Column Copier (Local)"

# ---------------------------
# Scrollable Frame helper
# ---------------------------

class ScrollableFrame(ttk.Frame):
    """
    A vertically scrollable frame (Canvas + inner Frame + Scrollbar).
    Put widgets into `self.scrollable_frame`.
    """
    def __init__(self, container, *args, **kwargs):
        super().__init__(container, *args, **kwargs)

        self.canvas = tk.Canvas(self, borderwidth=0, highlightthickness=0)
        self.vsb = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = ttk.Frame(self.canvas)

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        )

        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=self.vsb.set)

        self.canvas.pack(side="left", fill="both", expand=True)
        self.vsb.pack(side="right", fill="y")

        # Mouse wheel support
        self._bind_mousewheel(self.canvas)

    def _bind_mousewheel(self, widget):
        # Windows / macOS
        widget.bind_all("<MouseWheel>", self._on_mousewheel)
        # Linux (scroll up/down)
        widget.bind_all("<Button-4>", self._on_mousewheel_linux_up)
        widget.bind_all("<Button-5>", self._on_mousewheel_linux_down)

    def _on_mousewheel(self, event):
        self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def _on_mousewheel_linux_up(self, _event):
        self.canvas.yview_scroll(-3, "units")

    def _on_mousewheel_linux_down(self, _event):
        self.canvas.yview_scroll(3, "units")


# ---------------------------
# Normalization helpers (headers & keys)
# ---------------------------

def _norm_header(name: str) -> str:
    """Normalize a header: strip spaces, collapse internal whitespace, casefold."""
    if name is None:
        return ""
    return " ".join(str(name).strip().split()).casefold()

def _norm_headers_map(names):
    """Return dict original_name -> normalized_name."""
    return {n: _norm_header(n) for n in names}

def _norm_key(val):
    """Normalize a key for matching: coerce to string, strip, casefold (None for NaN/empty)."""
    if pd.isna(val):
        return None
    s = str(val).strip()
    return s.casefold() if s else None


# ---------------------------
# Excel helpers
# ---------------------------

def read_excel_sheets(path: str) -> Dict[str, pd.DataFrame]:
    """Read an Excel file into {sheet_name: DataFrame}."""
    xls = pd.ExcelFile(path)
    return {s: xls.parse(s) for s in xls.sheet_names}

def open_wb(path: Optional[str]) -> Workbook:
    """Open existing workbook or create a new blank one."""
    if path and os.path.exists(path):
        return load_workbook(path, data_only=False)
    return Workbook()

def get_ws(wb: Workbook, name: str) -> Worksheet:
    """Get or create worksheet by name."""
    if name in wb.sheetnames:
        return wb[name]
    return wb.create_sheet(title=name)

def header_map(ws: Worksheet) -> Dict[str, int]:
    """
    Normalized header name -> 1-based column index.
    Assumes headers in row 1. If duplicates (after normalization), keep the first.
    """
    headers = {}
    row = next(ws.iter_rows(min_row=1, max_row=1), [])
    for idx, cell in enumerate(row, start=1):
        val = cell.value
        if val is None:
            continue
        norm = _norm_header(val)
        if norm and norm not in headers:
            headers[norm] = idx
    return headers

def ensure_headers(ws: Worksheet, columns: List[str], existing_map: Optional[Dict[str, int]] = None) -> Dict[str, int]:
    """
    Ensure all columns exist as headers in row 1; stores mapping by *normalized* name.
    If a requested column exists with different spacing/case, we reuse it (no duplicate).
    """
    hmap = existing_map or header_map(ws)

    # If the sheet looks empty, write all headers left-to-right with the original labels
    if not hmap:
        for j, col in enumerate(columns, start=1):
            ws.cell(row=1, column=j, value=col)
        return header_map(ws)  # rebuild normalized map

    # Add missing headers at the rightmost position
    next_col = max(hmap.values()) + 1
    for col in columns:
        norm = _norm_header(col)
        if norm not in hmap:
            ws.cell(row=1, column=next_col, value=col)  # keep original label/case
            hmap[norm] = next_col
            next_col += 1
    return hmap

def detect_formula_cell(cell) -> bool:
    """Return True if cell contains a formula."""
    val = cell.value
    if isinstance(val, str) and val.startswith("="):
        return True
    if getattr(cell, "data_type", None) == "f":
        return True
    return False

def build_dest_key_row_map(ws: Worksheet, key_col_name: str, hmap: Dict[str, int]) -> Dict[object, int]:
    """
    Return mapping: normalized_key_value -> worksheet_row_index (1-based).
    Only rows with non-empty key are included.
    """
    norm_key_header = _norm_header(key_col_name)
    if norm_key_header not in hmap:
        return {}

    key_col = hmap[norm_key_header]
    mapping = {}
    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=key_col).value
        nval = _norm_key(val)
        if nval:
            mapping[nval] = r
    return mapping

def find_next_empty_row(ws: Worksheet, key_col_name: str, hmap: Dict[str, int]) -> int:
    """Find the next empty row by scanning the key column from the bottom up (more reliable than max_row)."""
    norm_key_header = _norm_header(key_col_name)
    if norm_key_header not in hmap:
        return 2  # first data row when headers just added
    key_col = hmap[norm_key_header]

    r = ws.max_row
    while r >= 2:
        val = ws.cell(row=r, column=key_col).value
        if val is not None and str(val).strip() != "":
            break
        r -= 1
    return r + 1 if r >= 1 else 2


# ---------------------------
# Core operations
# ---------------------------

def run_overwrite(
    wb: Workbook,
    ws_name: str,
    src_df: pd.DataFrame,
    src_key: str,
    dest_key: str,
    selected_cols: List[str],
    add_missing_ids: bool,
    allow_overwrite_formulas: bool,  # NEW: toggle formula overwrite
) -> Dict[str, int]:
    """
    Key-based overwrite with header/key normalization and optional formula overwrite.
    Returns counters: {"updated": X, "skipped_formulas": Y, "new_rows": Z}
    """
    ws = get_ws(wb, ws_name)
    required_headers = list({dest_key, *selected_cols})
    hmap = ensure_headers(ws, required_headers, header_map(ws))

    # destination row map by normalized key
    key_to_row = build_dest_key_row_map(ws, dest_key, hmap)

    updated = 0
    skipped = 0
    new_rows = 0

    if src_key not in src_df.columns:
        raise ValueError(f"Source key '{src_key}' not found in source sheet.")

    # normalize source key column once
    src_df = src_df.copy()
    src_df["_norm_key"] = src_df[src_key].apply(_norm_key)

    for _, srow in src_df.iterrows():
        nkey = srow["_norm_key"]
        if not nkey:
            continue

        dest_row = key_to_row.get(nkey)
        if dest_row is None:
            if add_missing_ids:
                # ensure headers are still valid (in case of late additions)
                hmap = ensure_headers(ws, required_headers, hmap)
                dest_row = find_next_empty_row(ws, dest_key, hmap)
                # write key
                ws.cell(row=dest_row, column=hmap[_norm_header(dest_key)], value=srow[src_key])
                # write selected columns
                for col in selected_cols:
                    ws.cell(row=dest_row, column=hmap[_norm_header(col)], value=srow.get(col, None))
                key_to_row[nkey] = dest_row
                new_rows += 1
            continue

        # Update existing row cells
        for col in selected_cols:
            col_idx = hmap.get(_norm_header(col))
            if not col_idx:
                # header unexpectedly missing; ensure and refresh map
                hmap = ensure_headers(ws, [col], hmap)
                col_idx = hmap[_norm_header(col)]
            cell = ws.cell(row=dest_row, column=col_idx)
            if detect_formula_cell(cell) and not allow_overwrite_formulas:
                skipped += 1
                continue
            cell.value = srow.get(col, None)
            updated += 1

    return {"updated": updated, "skipped_formulas": skipped, "new_rows": new_rows}


def run_append(
    wb: Workbook,
    ws_name: str,
    src_df: pd.DataFrame,
    src_key: str,
    dest_key: str,
    selected_cols: List[str],
) -> Dict[str, int]:
    """Append all source rows; writes key + selected columns to the bottom (safe next-row finder)."""
    ws = get_ws(wb, ws_name)
    required_headers = list({dest_key, *selected_cols})
    hmap = ensure_headers(ws, required_headers, header_map(ws))

    rows_added = 0
    next_row = find_next_empty_row(ws, dest_key, hmap)

    for _, srow in src_df.iterrows():
        # key
        ws.cell(row=next_row, column=hmap[_norm_header(dest_key)], value=srow.get(src_key, None))
        # selected cols
        for col in selected_cols:
            ws.cell(row=next_row, column=hmap[_norm_header(col)], value=srow.get(col, None))
        next_row += 1
        rows_added += 1

    return {"rows_added": rows_added}


# ---------------------------
# GUI
# ---------------------------

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("820x700")
        self.resizable(True, True)

        # Scrollable container
        container = ScrollableFrame(self)
        container.pack(fill="both", expand=True)
        self.main_frame = container.scrollable_frame  # parent for all UI sections

        # State
        self.src_path: Optional[str] = None
        self.dest_path: Optional[str] = None
        self.src_sheets: Dict[str, pd.DataFrame] = {}
        self.dest_sheets: Dict[str, pd.DataFrame] = {}

        # Build widgets inside scrollable frame
        self.create_widgets()

    def create_widgets(self):
        pad = {"padx": 8, "pady": 6}

        # 1) Files
        frm_files = ttk.LabelFrame(self.main_frame, text="1) Choose files")
        frm_files.pack(fill="x", **pad)

        ttk.Button(frm_files, text="Select Source .xlsx", command=self.pick_source).grid(row=0, column=0, sticky="w", **pad)
        self.lbl_src = ttk.Label(frm_files, text="No source selected")
        self.lbl_src.grid(row=0, column=1, sticky="w", **pad)

        ttk.Button(frm_files, text="Select Destination .xlsx", command=self.pick_dest).grid(row=1, column=0, sticky="w", **pad)
        self.lbl_dest = ttk.Label(frm_files, text="(Optional) No destination selected")
        self.lbl_dest.grid(row=1, column=1, sticky="w", **pad)

        # 2) Sheets
        frm_sheets = ttk.LabelFrame(self.main_frame, text="2) Choose sheets")
        frm_sheets.pack(fill="x", **pad)

        ttk.Label(frm_sheets, text="Source sheet:").grid(row=0, column=0, sticky="w", **pad)
        self.cmb_src_sheet = ttk.Combobox(frm_sheets, state="readonly")
        self.cmb_src_sheet.grid(row=0, column=1, sticky="we", **pad)
        self.cmb_src_sheet.bind("<<ComboboxSelected>>", self.on_src_sheet_change)

        ttk.Label(frm_sheets, text="Destination sheet:").grid(row=1, column=0, sticky="w", **pad)
        self.cmb_dest_sheet = ttk.Combobox(frm_sheets)  # editable to allow new names
        self.cmb_dest_sheet.grid(row=1, column=1, sticky="we", **pad)

        frm_sheets.columnconfigure(1, weight=1)

        # 3) Keys & columns
        frm_keys = ttk.LabelFrame(self.main_frame, text="3) Keys & columns")
        frm_keys.pack(fill="both", expand=True, **pad)

        ttk.Label(frm_keys, text="Source key:").grid(row=0, column=0, sticky="w", **pad)
        self.cmb_src_key = ttk.Combobox(frm_keys, state="readonly")
        self.cmb_src_key.grid(row=0, column=1, sticky="we", **pad)

        ttk.Label(frm_keys, text="Destination key:").grid(row=1, column=0, sticky="w", **pad)
        self.cmb_dest_key = ttk.Combobox(frm_keys)  # editable
        self.cmb_dest_key.grid(row=1, column=1, sticky="we", **pad)

        ttk.Label(frm_keys, text="Columns to copy from source:").grid(row=2, column=0, sticky="nw", **pad)
        self.lst_columns = tk.Listbox(frm_keys, selectmode=tk.MULTIPLE, exportselection=False, height=10)
        self.lst_columns.grid(row=2, column=1, sticky="nsew", **pad)

        frm_keys.rowconfigure(2, weight=1)
        frm_keys.columnconfigure(1, weight=1)

        # 4) Mode
        frm_mode = ttk.LabelFrame(self.main_frame, text="4) Mode")
        frm_mode.pack(fill="x", **pad)

        self.mode = tk.StringVar(value="overwrite")
        ttk.Radiobutton(frm_mode, text="Overwrite (key-based)", variable=self.mode, value="overwrite").grid(row=0, column=0, sticky="w", **pad)
        ttk.Radiobutton(frm_mode, text="Append (add rows to bottom)", variable=self.mode, value="append").grid(row=0, column=1, sticky="w", **pad)

        self.var_add_missing = tk.BooleanVar(value=False)
        ttk.Checkbutton(frm_mode, text="Also add rows for source IDs not in destination (overwrite mode only)", variable=self.var_add_missing).grid(row=1, column=0, columnspan=2, sticky="w", **pad)

        self.var_overwrite_file = tk.BooleanVar(value=False)
        ttk.Checkbutton(frm_mode, text="Overwrite destination file on save (otherwise 'Save As...')", variable=self.var_overwrite_file).grid(row=2, column=0, columnspan=2, sticky="w", **pad)

        # NEW: allow overwriting formulas toggle
        self.var_allow_formula_overwrite = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            frm_mode,
            text="Allow overwriting formulas in selected columns",
            variable=self.var_allow_formula_overwrite
        ).grid(row=3, column=0, columnspan=2, sticky="w", **pad)

        # Actions
        frm_action = ttk.Frame(self.main_frame)
        frm_action.pack(fill="x", **pad)
        ttk.Button(frm_action, text="Run", command=self.run_operation).pack(side="left", padx=4)
        ttk.Button(frm_action, text="Quit", command=self.destroy).pack(side="right", padx=4)

        # Log
        frm_log = ttk.LabelFrame(self.main_frame, text="Log")
        frm_log.pack(fill="both", expand=True, **pad)
        self.txt_log = tk.Text(frm_log, height=10)
        self.txt_log.pack(fill="both", expand=True, padx=6, pady=6)

    # ---------------------------
    # UI handlers
    # ---------------------------

    def pick_source(self):
        path = filedialog.askopenfilename(title="Select Source Excel", filetypes=[("Excel files", "*.xlsx")])
        if not path:
            return
        try:
            self.src_sheets = read_excel_sheets(path)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to read source: {e}")
            return
        self.src_path = path
        self.lbl_src.config(text=path)

        # Populate source sheet names & default selection
        self.cmb_src_sheet["values"] = list(self.src_sheets.keys())
        if self.cmb_src_sheet["values"]:
            self.cmb_src_sheet.current(0)
            self.on_src_sheet_change()

        # If destination sheet name is empty, propose same as source
        if not self.cmb_dest_sheet.get():
            self.cmb_dest_sheet.set(self.cmb_src_sheet.get() or "Sheet1")

    def pick_dest(self):
        path = filedialog.askopenfilename(title="Select Destination Excel (optional)", filetypes=[("Excel files", "*.xlsx")])
        if not path:
            return
        try:
            self.dest_sheets = read_excel_sheets(path)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to read destination: {e}")
            return
        self.dest_path = path
        self.lbl_dest.config(text=path)

        # Populate destination sheet names (allow typing new)
        self.cmb_dest_sheet["values"] = list(self.dest_sheets.keys())
        if self.cmb_dest_sheet["values"] and not self.cmb_dest_sheet.get():
            self.cmb_dest_sheet.current(0)

        # Populate dest key combo if we can infer columns (from current dest sheet selection)
        if self.cmb_dest_sheet.get():
            df = self.dest_sheets.get(self.cmb_dest_sheet.get(), pd.DataFrame())
            self.cmb_dest_key["values"] = list(df.columns)
        else:
            self.cmb_dest_key["values"] = []

    def on_src_sheet_change(self, *_):
        sname = self.cmb_src_sheet.get()
        df = self.src_sheets.get(sname, pd.DataFrame())
        cols = list(df.columns)

        # Populate key and columns controls
        self.cmb_src_key["values"] = cols
        if "ID" in cols:
            self.cmb_src_key.set("ID")
        elif cols:
            self.cmb_src_key.current(0)

        # Fill listbox with all columns (user can exclude key later)
        self.lst_columns.delete(0, tk.END)
        for c in cols:
            self.lst_columns.insert(tk.END, c)

        # Populate destination key suggestions from currently selected destination sheet
        dname = self.cmb_dest_sheet.get()
        ddf = self.dest_sheets.get(dname, pd.DataFrame())
        self.cmb_dest_key["values"] = list(ddf.columns)

        # Default dest key = same as source key if present, else "ID"
        src_key = self.cmb_src_key.get() or "ID"
        if src_key in self.cmb_dest_key["values"]:
            self.cmb_dest_key.set(src_key)
        elif "ID" in self.cmb_dest_key["values"]:
            self.cmb_dest_key.set("ID")

    # ---------------------------
    # Execution
    # ---------------------------

    def run_operation(self):
        # Validations
        if not self.src_path:
            messagebox.showwarning("Missing source", "Please select a Source Excel file.")
            return

        src_sheet = self.cmb_src_sheet.get()
        if not src_sheet:
            messagebox.showwarning("Missing sheet", "Please choose a Source sheet.")
            return

        dest_sheet = self.cmb_dest_sheet.get().strip() or "Sheet1"

        # Source DF
        src_df = self.src_sheets.get(src_sheet, pd.DataFrame())
        if src_df.empty:
            messagebox.showerror("Empty source", "Selected Source sheet is empty.")
            return

        src_key = (self.cmb_src_key.get() or "").strip()
        if not src_key or src_key not in src_df.columns:
            messagebox.showerror("Key required", "Please choose a valid Source key column (e.g., 'ID').")
            return

        # Columns to copy
        sel_indices = self.lst_columns.curselection()
        if not sel_indices:
            messagebox.showwarning("No columns", "Select at least one column from the list.")
            return
        selected_cols = [self.lst_columns.get(i) for i in sel_indices]
        # Don’t copy the key itself
        selected_cols = [c for c in selected_cols if c != src_key]

        # Destination key
        dest_key = (self.cmb_dest_key.get() or "").strip() or src_key

        # Load or create workbook
        try:
            wb = open_wb(self.dest_path)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open destination workbook: {e}")
            return

        # Run
        try:
            if self.mode.get() == "overwrite":
                counters = run_overwrite(
                    wb=wb,
                    ws_name=dest_sheet,
                    src_df=src_df,
                    src_key=src_key,
                    dest_key=dest_key,
                    selected_cols=selected_cols,
                    add_missing_ids=self.var_add_missing.get(),
                    allow_overwrite_formulas=self.var_allow_formula_overwrite.get(),  # NEW
                )
                log = (
                    f"Overwrite done. Updated cells: {counters['updated']}, "
                    f"Skipped formula cells: {counters['skipped_formulas']}, "
                    f"New rows: {counters['new_rows']}.\n"
                )
            else:
                counters = run_append(
                    wb=wb,
                    ws_name=dest_sheet,
                    src_df=src_df,
                    src_key=src_key,
                    dest_key=dest_key,
                    selected_cols=selected_cols,
                )
                log = f"Append done. Rows added: {counters['rows_added']}.\n"

            # Save
            if self.dest_path and self.var_overwrite_file.get():
                save_path = self.dest_path
            else:
                default_name = "updated_destination.xlsx"
                save_path = filedialog.asksaveasfilename(
                    title="Save As",
                    defaultextension=".xlsx",
                    initialfile=default_name,
                    filetypes=[("Excel files", "*.xlsx")],
                )
                if not save_path:
                    self.append_log("Save cancelled.\n")
                    return

            wb.save(save_path)
            self.append_log(log + f"Saved to: {save_path}\n")
            messagebox.showinfo("Success", f"Operation completed.\nSaved to:\n{save_path}")

        except Exception as e:
            messagebox.showerror("Error", f"Operation failed: {e}")
            self.append_log(f"Error: {e}\n")

    def append_log(self, text: str):
        self.txt_log.insert(tk.END, text)
        self.txt_log.see(tk.END)


# ---------------------------
# Entry point
# ---------------------------

def main():
    app = App()
    app.mainloop()

if __name__ == "__main__":
    main()